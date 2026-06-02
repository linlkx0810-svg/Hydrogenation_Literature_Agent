"""
excel_io.py
Styled Excel export helpers used by all pipeline modules.
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)


def _style_sheet(sheet, columns: list[str], n_rows: int) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx, col in enumerate(columns, start=1):
        lengths = [
            len(str(sheet.cell(r, idx).value or ""))
            for r in range(1, min(n_rows + 2, 120))
        ]
        width = min(max(max(lengths, default=len(col)) + 2, 12), 52)
        sheet.column_dimensions[sheet.cell(1, idx).column_letter].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if n_rows > 0:
        tname = re.sub(r"[^A-Za-z0-9]", "", sheet.title.title()) + "Table"
        table = Table(displayName=tname, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        sheet.add_table(table)


def export_records_xlsx(
    records: list[dict],
    columns: list[str],
    output_path: Path,
    sheet_name: str = "records",
    extra_sheets: dict[str, list[dict]] | None = None,
    extra_sheet_columns: dict[str, list[str]] | None = None,
) -> None:
    """
    Write records to a styled Excel workbook.

    Parameters
    ----------
    records      : list of dicts (main sheet)
    columns      : column order / names
    output_path  : destination .xlsx path
    sheet_name   : name for the main sheet
    extra_sheets : optional {sheet_name: [records]} for additional sheets
    extra_sheet_columns : optional {sheet_name: [columns]} overrides for extra sheets
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    def _add(name: str, rows: list[dict], sheet_columns: list[str]) -> None:
        ws = wb.create_sheet(name)
        ws.append(sheet_columns)
        for row in rows:
            ws.append([str(row.get(c, "") or "") for c in sheet_columns])
        _style_sheet(ws, sheet_columns, len(rows))

    _add(sheet_name, records, columns)
    for name, rows in (extra_sheets or {}).items():
        sheet_columns = (extra_sheet_columns or {}).get(name, columns)
        _add(name, rows, sheet_columns)

    wb.save(output_path)


def append_log_xlsx(log_row: dict, columns: list[str], log_path: Path) -> None:
    """
    Append a single log row to an Excel file (creates the file if absent).
    Deduplication by Row index is applied if 'Row' is in columns.
    """
    log_path = Path(log_path)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    new_df = pd.DataFrame([{c: log_row.get(c, "") for c in columns}])

    if log_path.exists():
        existing = pd.read_excel(log_path)
        if "Row" in columns and "Row" in existing.columns:
            existing = existing[existing["Row"] != log_row.get("Row")]
        combined = pd.concat([existing, new_df], ignore_index=True)
        if "Row" in combined.columns:
            combined.sort_values("Row", inplace=True)
    else:
        combined = new_df

    combined.to_excel(log_path, index=False)
